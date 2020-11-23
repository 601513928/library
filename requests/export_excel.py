import xlrd
import xlwt

from openpyxl.workbook import Workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.utils import get_column_letter
from openpyxl.reader.excel import load_workbook

class HandleExcel():
    '''Excel 相关操作类'''
    # def __init__(self):
    #     self.head_row_labels = [u'学生ID', u'学生姓名', u'联系方式', u'知识点ID',  ]
    def write_to_excel_with_openpyxl(self, records, head_row, save_excel_name="save.xlsx"):
        #新建一个workbook
        wb = Workbook()
        #新建一个excelWriter
        ew = ExcelWriter(workbook=wb)
        #设置文件输出路径与名称
        dest_filename = save_excel_name.decode('utf-8')
        #第一个sheet是ws
        ws = wb.worksheets[0]
        #设置ws的名称
        ws.title = "range names"
        #写第一行，标题行
        for h_x in range(1, len(head_row) + 1):
            h_col = get_column_letter(h_x)
            #print h_col
            ws.cell('%s%s' % (h_col, 1)).value = '%s' % (head_row[h_x - 1])
        
        i = 2
        for record in records:
            record_list = str(record).strip().split("\t")
            for x in range(1, len(record_list) + 1):
                col = get_column_letter(x)
                ws.cell('%s%s' % (col, i)).value = '%s' % (record_list[x - 1].decode('utf-8'))
            i += 1
        ew.save(filename=dest_filename)
